import cx_Oracle
from openpyxl import load_workbook
import os
import re

# primary data
table_path = input('Input table path:')[1:-1]  # substring for drag and drop into console
id_form = input('Input code parent form:')
file_name = os.path.splitext(os.path.basename(table_path))[0]

'''
input example
C:\\Users\ARM2\Desktop\Исследование функции внешнего дыхания.xlsx
11222242
'''


def check_element(element='0'):
    if element:
        element = element.lower()
        if re.search(r'заг.ловок', element):
            element = 0
        elif re.search(r'вопрос', element):
            element = 1
    else:
        element = 1
    return element


def check_type_answer(type_answer):
    if type_answer:
        type_answer = type_answer.lower()
        if re.search(r'свободный ответ', type_answer) or re.search(r'простой текст', type_answer):
            type_answer = 0
        elif re.search(r'значени[ея] из списка', type_answer):
            type_answer = 2
    else:
        type_answer = 0
    return type_answer


def check_multi_choise(multi_choise):
    if multi_choise:
        multi_choise.lower()
        if re.search(r'да', multi_choise) or re.search(r'/+', multi_choise):
            multi_choise = 1
        elif re.search(r'н.т', multi_choise) or re.search(r'-', multi_choise):
            multi_choise = 0
        else:
            multi_choise = 0
    else:
        multi_choise = 0
    return multi_choise


def check_answers(answers):
    if ws.cell(excel_row, 5).value:
        answers = ws.cell(excel_row, 5).value.split('\n')
        if isinstance(answers[0], str):
            answers = re.split(r'\s{2,}', ws.cell(excel_row, 5).value)
    elif ws.cell(excel_row, 5).value == None:
        answers = ['']
    return answers

def check_type_element_data(type_element_data):
    type_element_data = str(type_element_data)
    if re.search('дата', type_element_data.lower()):
        type_element_data = 1
    if re.search('время', type_element_data.lower()):
        type_element_data = 2
    if re.search('дата и время', type_element_data.lower()):
        type_element_data = 3
    if re.search('логический', type_element_data.lower()):
        type_element_data = 4
    else:
        type_element_data = 0
    return type_element_data
# get data from excel
wb = load_workbook(table_path)
protocols = {}

for ws in wb.worksheets:
    sheet_name = ''
    for i in range(1, 3):  # check name protocol in the first three excel_row
        if ws.cell(i, 1).value:
            sheet_name = ws.cell(i, 1).value
            break
    protocol_rows = []
    for excel_row in range(5, ws.max_row):  # 5 excel_row is begin protocol_row
        element_data = ws.cell(excel_row, 2).value
        type_element_data = check_type_element_data(element_data) if element_data else False
        if element_data == None:
            if ws.cell(excel_row, 5).value:
                for i in check_answers(ws.cell(excel_row, 5).value):
                    protocol_rows[len(protocol_rows) - 1][4].append(i)
                continue
            else:
                continue



        element = check_element(ws.cell(excel_row, 1).value)
        # 0 - разделитель
        # 1 - редактируемый
        # 2 - невидимый
        # 3 - только
        # чтение
        # 4 - неактивный;
        type_answer = check_type_answer(ws.cell(excel_row, 3).value)
        # 0 - Простой текст (строка) 1 - Описание (много строк) 2 - Значения из списка
        # 3 - Значения из внешнего справочника 4 - Формула 5 - Таблица 6 - Значение из системного справочника
        # 7 - Формула SQL 8 - Поле для ввода диагноза 9 - Поле для ввода услуг 10 - Значение из дерева
        # 11 - Значение из списка отмеченное галочками 12 - Файл 13 - Схема (изображение);
        multi_choise = check_multi_choise(ws.cell(excel_row, 4).value)
        # 0 or 1
        answers = check_answers(ws.cell(excel_row, 5).value)
        # Нарушений легочной вентиляции не зарегистрировано
        # Проба с физической нагрузкой - положительная
        excel_row = [element, element_data, type_answer, multi_choise, answers, type_element_data]
        protocol_rows.append(excel_row)
    protocols[sheet_name] = protocol_rows

# operation with oracle database
connection = cx_Oracle.connect('solution_med/elsoft@med')
print(connection.version)
create_form = connection.cursor()
parent_id = int(connection.cursor().execute("select id from SOLUTION_FORM.FORM where "
                                            "code ='{id_form}' and rownum = 1".format(id_form=id_form)).fetchone()[0])

for protocol_name, protocol_value in protocols.items():
    try:
        code_form = int(connection.cursor().execute("SELECT MAX(TO_NUMBER(code)) + 1 FROM solution_form.form where"
                                                    " trim(TRANSLATE(code, '0123456789-,.', ' ')) is null").fetchone()[
                            0])
        add_form = "DECLARE rc pkg_global.ref_cursor_type;" \
                   " BEGIN p_content.save_form(" \
                   "NULL, NULL, {parent_id}, {id_form}," \
                   " {id_form},  '{protocol_name}', ''," \
                   " 0.0, 1, 1, 0," \
                   " '', NULL, NULL, 0, 0, rc);COMMIT;END;".format(protocol_name=str(protocol_name),
                                                                   id_form=code_form, parent_id=parent_id)
        create_form.execute(add_form)
        id = int(connection.cursor().execute("select id from SOLUTION_FORM.FORM where "
                                             "code = '{code_form}'"
                                             " and rownum = 1".format(code_form=code_form)).fetchone()[0])
    except Exception as e:
        print(str(protocol_name))
        print('----------------------------------------------------------------------------------')
        print(e)
        print('----------------------------------------------------------------------------------')
        input("Print enter to exist")
        continue

    for index, item_name in enumerate(protocol_value):
        try:
            insert_form_item = """
            DECLARE rc pkg_global.ref_cursor_type;
            BEGIN
            p_content.save_form_item(
                NULL
                , NULL
                , '{id}'
                , NULL
                , NULL
                , '{code}'
                , {sortcode}
                , {type}
                , {type_value}
                , NULL
                , 0.0
                , '{item_name}'
                , NULL
                , 1
                , ''
                , {type_value}
                , {type_element_data}
                , {is_multi}
                , NULL
                , NULL
                , NULL
                , ''
                , ''
                , 0
                , 0
                , 0
                , ''
                , 0
                , NULL
                , ''
                , 0
                , NULL
                , rc);
                COMMIT;
                END;
                """.format(item_name=item_name[1],
                           id=id, code=index,
                           sortcode=index,
                           type=item_name[0],
                           type_value=item_name[2],
                           is_multi=item_name[3],
                           type_element_data=item_name[5])
            create_form.execute(insert_form_item)
            get_if_form_item = int(
                connection.cursor().execute("SELECT SOLUTION_MED.PKG_GLOBAL.GET_NEXT_ID('SOLUTION_FORM',"
                                            " 'FORM_ITEM') - 1 FROM DUAL").fetchone()[0])
        except Exception as e:
            print(str(protocol_name) + ' :' + str(item_name))
            print('----------------------------------------------------------------------------------')
            print(e)
            print('----------------------------------------------------------------------------------')
            input("Press enter to continue")
            continue

        if item_name[2] == 2:

                for index, answer in enumerate(item_name[4]):
                    try:
                        get_if_form_item_value = int(
                            connection.cursor().execute("SELECT SOLUTION_MED.PKG_GLOBAL.GET_NEXT_ID('SOLUTION_FORM',"
                                                        " 'FORM_ITEM_VALUE') FROM DUAL").fetchone()[0])
                        insert_form_item_value = """
                        INSERT INTO SOLUTION_FORM.FORM_ITEM_VALUE (
                        ID
                        ,FORM_ITEM_ID
                        ,CODE
                        ,SORTCODE
                        ,TEXT
                        ,NOTE
                        ,BALL
                        ,STATUS
                        ,IS_DEFAULT
                        ,NAME
                        ,ROOT_ID
                        ,IGNORE_TEXT
                        ) VALUES (
                        '{form_item_value}'
                        ,'{form_item}'
                        ,'{code}'
                        , {sort_code}
                        ,'{answer}'
                        ,''
                        ,NULL
                        ,1
                        ,0
                        ,'{name_answer}'
                        ,''
                        ,NULL
                        )
                        """.format(form_item=get_if_form_item,
                                   form_item_value=get_if_form_item_value,
                                   answer=answer, name_answer=answer, code=index, sort_code=index)
                        create_form.execute(insert_form_item_value)
                        create_form.execute('COMMIT')
                    except Exception as e:
                        print(str(protocol_name) + ' : ' + str(item_name) + ' : ' + str(answer))
                        print('----------------------------------------------------------------------------------')
                        print(e)
                        print('----------------------------------------------------------------------------------')
                        input("Print enter to exist")
                        continue

